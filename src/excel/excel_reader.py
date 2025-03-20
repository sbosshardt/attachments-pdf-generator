#!/usr/bin/env python3
"""
Excel reader module to read attachment data from Excel files.
"""

import os
import openpyxl
from src.config.paths import EXCEL_FILE, SHEET_NAME

def read_attachment_data(for_toc=True):
    """
    Read attachment data from Excel file.
    
    Args:
        for_toc (bool): Whether this is for TOC generation (True) or PDF merging (False)
                        Determines which fields are required.
    
    Returns:
        list: List of dictionaries containing attachment data
    """
    print(f"Opening Excel file: {EXCEL_FILE}")
    
    # Check if input file exists
    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(f"Excel file not found: {EXCEL_FILE}")
    
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    
    # Check if sheet exists
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found in {EXCEL_FILE}")
    
    sheet = workbook[SHEET_NAME]
    
    # Find headers
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value)
    
    # Field mapping (convert sheet headers to our standardized field names)
    field_mapping = {
        'Attachment Number': ['Attachment Number', 'Attachment #', 'Number'],
        'Title': ['Title', 'Document Title'],
        'Page count': ['Page count', 'Pages', 'Page Count'],
        'Additional Remarks about File': ['Additional Remarks about File', 'Remarks', 'Notes'],
        'Body': ['Body', 'Description', 'Body (Description)'],
        'Filename Reference': ['Filename Reference', 'Filename', 'File'],
        'Date': ['Date', 'Date (time Pacific)', 'Document Date'],
        'Language': ['Language', 'Lang', 'Language Code'],
        'Exclude': ['Exclude', 'Skip']
    }
    
    # Map headers to indices
    header_indices = {}
    for field, possible_headers in field_mapping.items():
        for i, header in enumerate(headers):
            if header and any(possible_match.lower() == header.lower() for possible_match in possible_headers):
                header_indices[field] = i
                break
    
    # Check if we found all required fields
    if for_toc:
        required_fields = ['Attachment Number', 'Title']
    else:
        required_fields = ['Attachment Number', 'Filename Reference']
        
    for field in required_fields:
        if field not in header_indices:
            raise ValueError(f"Required field '{field}' not found in Excel headers")
    
    data = []
    # Start from row 2 (skip header)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Skip empty rows
        if not any(row):
            continue
        
        # Skip rows marked as excluded
        exclude_idx = header_indices.get('Exclude')
        if exclude_idx is not None and row[exclude_idx] in (True, 'TRUE', 'True', 'true', 'YES', 'Yes', 'yes', '1', 1):
            continue
        
        attachment = {}
        for field, idx in header_indices.items():
            if field != 'Exclude':  # We've already used this field for filtering
                attachment[field] = row[idx] if idx < len(row) else None
        
        # Set default value for Language if not found
        if 'Language' not in attachment or not attachment['Language']:
            attachment['Language'] = 'EN'
        
        data.append(attachment)
    
    print(f"Found {len(data)} attachments")
    return data

def normalize_attachment_number(attachment_num):
    """
    Normalize attachment number to string representation.
    
    Args:
        attachment_num: The attachment number which could be a float, int, or string
        
    Returns:
        str: Normalized attachment number as string
    """
    if isinstance(attachment_num, float) and attachment_num.is_integer():
        return str(int(attachment_num))
    return str(attachment_num)

def normalize_page_count(page_count):
    """
    Normalize page count to integer.
    
    Args:
        page_count: The page count which could be a float, int, or string
        
    Returns:
        int: Normalized page count as integer
    """
    if isinstance(page_count, float) and page_count.is_integer():
        return int(page_count)
    elif not isinstance(page_count, int):
        try:
            return int(float(page_count))
        except (ValueError, TypeError):
            return 1
    return page_count

def load_attachments_from_excel():
    """
    Load attachments from Excel file and prepare them for PDF merging.
    This function normalizes attachment numbers and adds file paths.
    
    Returns:
        list: List of dictionaries containing prepared attachment data
    """
    # Read raw data from Excel
    attachments = read_attachment_data(for_toc=False)
    
    # Process each attachment
    processed_attachments = []
    for attachment in attachments:
        # Normalize attachment number
        attachment_num = normalize_attachment_number(attachment.get('Attachment Number', ''))
        
        # Create a new dictionary with processed data
        processed = {
            'Number': attachment_num,
            'Title': attachment.get('Title', f'Attachment {attachment_num}'),
            'Language': attachment.get('Language', 'EN'),
        }
        
        # Add file path
        filename = attachment.get('Filename Reference', '')
        if filename:
            language = processed['Language'].lower()
            filepath = os.path.join('input-files', language, filename)
            processed['FilePath'] = filepath
            
            # Check if file exists
            if not os.path.exists(filepath):
                print(f"Warning: File for Attachment {attachment_num} not found: {filepath}")
        
        processed_attachments.append(processed)
    
    # Sort by attachment number
    processed_attachments.sort(key=lambda x: float(x['Number']) if x['Number'].replace('.', '', 1).isdigit() else float('inf'))
    
    return processed_attachments 