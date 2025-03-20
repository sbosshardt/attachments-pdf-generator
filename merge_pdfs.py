#!/usr/bin/env python3
"""
Merge PDF Attachments

This script merges the table of contents and cover pages PDF with the actual PDF attachments.
"""

import os
import openpyxl
import fitz  # PyMuPDF

# File paths
EXCEL_FILE = os.path.join('input-files', 'input-pdfs.xlsx')
TOC_PDF = os.path.join('output-files', 'weasyoutput.pdf')
MERGED_PDF = os.path.join('output-files', 'merged-attachments.pdf')

def read_attachment_data():
    """
    Read attachment data from Excel file.
    
    Returns:
        list: List of dictionaries containing attachment data
    """
    print(f"Opening Excel file: {EXCEL_FILE}")
    
    # Check if input file exists
    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(f"Excel file not found: {EXCEL_FILE}")
    
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    
    # Check if sheet exists
    if "Attachments Prep" not in workbook.sheetnames:
        raise ValueError(f"Sheet 'Attachments Prep' not found in {EXCEL_FILE}")
    
    sheet = workbook["Attachments Prep"]
    
    # Find headers
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value)
    
    # Field mapping (convert sheet headers to our standardized field names)
    field_mapping = {
        'Attachment Number': ['Attachment Number', 'Attachment #', 'Number'],
        'Title': ['Title', 'Document Title'],
        'Filename Reference': ['Filename Reference', 'Filename', 'File'],
        'Language': ['Language', 'Lang', 'Language Code'],
        'Exclude': ['Exclude', 'Skip']
    }
    
    # Map headers to indices
    header_indices = {}
    for field, possible_headers in field_mapping.items():
        for i, header in enumerate(headers):
            if header and any(possible_match.lower() == header.lower() for possible_match in possible_headers):
                header_indices[field] = i
                break
    
    # Check if we found all required fields
    required_fields = ['Attachment Number', 'Filename Reference']
    for field in required_fields:
        if field not in header_indices:
            raise ValueError(f"Required field '{field}' not found in Excel headers")
    
    data = []
    # Start from row 2 (skip header)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Skip empty rows
        if not any(row):
            continue
        
        # Skip rows marked as excluded
        exclude_idx = header_indices.get('Exclude')
        if exclude_idx is not None and row[exclude_idx] in (True, 'TRUE', 'True', 'true', 'YES', 'Yes', 'yes', '1', 1):
            continue
        
        attachment = {}
        for field, idx in header_indices.items():
            if field != 'Exclude':  # We've already used this field for filtering
                attachment[field] = row[idx] if idx < len(row) else None
        
        # Set default value for Language if not found
        if 'Language' not in attachment or not attachment['Language']:
            attachment['Language'] = 'EN'
            
        data.append(attachment)
    
    print(f"Found {len(data)} attachments")
    return data

def merge_pdfs(attachments):
    """
    Merge the TOC PDF with the attachment PDFs, placing each attachment after its cover page.
    
    Args:
        attachments (list): List of attachment data dictionaries
    """
    print(f"Merging PDFs into: {MERGED_PDF}")
    
    # Check if TOC PDF exists
    if not os.path.exists(TOC_PDF):
        raise FileNotFoundError(f"Table of contents PDF not found: {TOC_PDF}")
    
    # Load the TOC and cover pages PDF
    toc_pdf = fitz.open(TOC_PDF)
    print(f"TOC PDF has {len(toc_pdf)} pages")
    
    # Create a new document for the final merged PDF
    merged_pdf = fitz.open()
    
    # Build a mapping of attachment numbers to their data
    attachment_map = {}
    for attachment in attachments:
        attachment_num = attachment.get('Attachment Number', '')
        if isinstance(attachment_num, float) and attachment_num.is_integer():
            attachment_num = int(attachment_num)
        attachment_map[str(attachment_num)] = attachment
    
    # First, scan the TOC PDF to locate all cover pages and links
    cover_page_indices = {}
    toc_page_index = -1
    toc_links = {}
    
    print("Scanning TOC PDF to locate cover pages and links...")
    
    # First locate the TOC page
    for page_num in range(toc_pdf.page_count):
        page = toc_pdf[page_num]
        text = page.get_text()
        if "Table of Contents" in text:
            toc_page_index = page_num
            print(f"DEBUG: Found TOC on page {page_num+1}")
            # Extract all links from TOC
            links = page.get_links()
            for link in links:
                if 'uri' in link and link['uri'].startswith('#cover-'):
                    attachment_id = link['uri'][7:]  # Remove '#cover-'
                    toc_links[attachment_id] = link
                    print(f"DEBUG: Found TOC link to Attachment {attachment_id}")
            break
    
    print(f"DEBUG: Found {len(toc_links)} TOC links")
    
    # Next, identify all cover pages by text pattern
    for page_num in range(toc_pdf.page_count):
        page = toc_pdf[page_num]
        text = page.get_text()
        
        if "Table of Contents" in text:
            continue  # Skip TOC page
            
        for attachment_id in attachment_map:
            if f"Attachment {attachment_id}" in text:
                # Make sure this isn't just a mention in another cover page
                if "Page" in text and f"Attachment {attachment_id}" in text.split('\n')[:5]:
                    cover_page_indices[attachment_id] = page_num
                    print(f"DEBUG: Found cover page for Attachment {attachment_id} on page {page_num+1}")
                    break
    
    print(f"DEBUG: Found {len(cover_page_indices)} cover pages")
    
    # Now process the TOC PDF
    # Keep track of mapping from original page to new page
    page_mapping = {}
    current_page = 0
    
    # Add all the TOC pages first
    for i in range(toc_pdf.page_count):
        merged_pdf.insert_pdf(toc_pdf, from_page=i, to_page=i)
        page_mapping[i] = current_page
        current_page += 1
    
    # Track positions for bookmarks and links
    bookmark_positions = {}
    
    # Now add each attachment right after its cover page
    for attachment_num, page_idx in sorted(cover_page_indices.items(), key=lambda x: int(x[0])):
        attachment = attachment_map.get(attachment_num)
        if not attachment:
            print(f"Warning: No data found for Attachment {attachment_num}, skipping")
            continue
            
        filename = attachment.get('Filename Reference', '')
        language = attachment.get('Language', 'EN')
        
        # Skip if no filename
        if not filename:
            print(f"Warning: No filename for Attachment {attachment_num}, skipping")
            continue
        
        # We already added the cover page in the first pass
        # Find its position in the merged PDF
        cover_page_pos = page_mapping[page_idx]
        bookmark_positions[attachment_num] = cover_page_pos
        
        # Create the full path to the attachment file
        attachment_path = os.path.join('input-files', language.lower(), filename)
        
        # Check if attachment exists
        if not os.path.exists(attachment_path):
            print(f"Warning: Attachment file not found: {attachment_path}, skipping")
            continue
        
        try:
            # Calculate where to insert
            insert_pos = cover_page_pos + 1
            
            # Open the attachment PDF
            attachment_pdf = fitz.open(attachment_path)
            attachment_pages = len(attachment_pdf)
            
            print(f"DEBUG: Inserting {attachment_pages} pages for Attachment {attachment_num} after position {insert_pos}")
            
            # Insert after the cover page
            merged_pdf.insert_pdf(attachment_pdf, start_at=insert_pos)
            
            # Update page mapping for all subsequent pages
            # This is critical for maintaining the correct link targets
            for orig_page in sorted(page_mapping.keys()):
                if page_mapping[orig_page] >= insert_pos:
                    page_mapping[orig_page] += attachment_pages
            
            # Also update bookmark positions for all subsequent attachments
            for att_num in bookmark_positions:
                if bookmark_positions[att_num] >= insert_pos:
                    bookmark_positions[att_num] += attachment_pages
            
            # Update current page counter
            current_page += attachment_pages
            
            print(f"Added: Attachment {attachment_num} - {filename} ({attachment_pages} pages)")
            
        except Exception as e:
            print(f"Error adding attachment {attachment_num}: {e}")
    
    # Create bookmarks/outline for the PDF
    toc = []
    
    # First add TOC bookmark if we found the TOC page
    if toc_page_index >= 0 and toc_page_index in page_mapping:
        toc.append([1, "Table of Contents", page_mapping[toc_page_index]])
    
    # Add bookmarks for each attachment that we found a position for
    for attachment_num in sorted(bookmark_positions.keys(), key=lambda x: int(x)):
        attachment = attachment_map.get(attachment_num)
        if not attachment:
            continue
            
        title = attachment.get('Title', 'Untitled')
        
        # Add bookmark for this attachment's cover page
        toc.append([1, f"Attachment {attachment_num}: {title}", bookmark_positions[attachment_num]])
    
    # Set the table of contents/bookmarks
    if toc:
        print(f"Setting {len(toc)} bookmarks in PDF")
        try:
            merged_pdf.set_toc(toc)
        except Exception as e:
            print(f"Warning: Failed to set bookmarks: {e}")
    
    # Fix links - we need to update all links in the TOC to point to the correct pages
    links_fixed = 0
    
    try:
        # First, focus on the TOC page for fixing links
        toc_page_pos = -1
        for i in range(merged_pdf.page_count):
            text = merged_pdf[i].get_text()
            if "Table of Contents" in text:
                toc_page_pos = i
                break
        
        if toc_page_pos >= 0:
            print(f"Fixing links on TOC page (page {toc_page_pos+1})")
            toc_page = merged_pdf[toc_page_pos]
            links = toc_page.get_links()
            print(f"Found {len(links)} links on TOC page")
            
            # Fix each link
            for link in links:
                if 'uri' in link and link['uri'].startswith('#cover-'):
                    attachment_id = link['uri'][7:]  # Remove '#cover-'
                    
                    # If we know the position of this attachment's cover page
                    if attachment_id in bookmark_positions:
                        target_page = bookmark_positions[attachment_id]
                        
                        # Create a new goto link
                        new_link = {
                            'kind': fitz.LINK_GOTO,
                            'from': link['from'],
                            'page': target_page,
                            'to': fitz.Point(0, 0),
                            'zoom': 0
                        }
                        
                        # Remove old link and add new one
                        toc_page.delete_link(link)
                        toc_page.insert_link(new_link)
                        links_fixed += 1
        
        # Now check all pages for any other internal links that need fixing
        for page_num in range(merged_pdf.page_count):
            if page_num == toc_page_pos:
                continue  # Skip TOC page as we've already processed it
                
            page = merged_pdf[page_num]
            links = page.get_links()
            
            for link in links:
                # If this is a named destination link
                if 'uri' in link and link['uri'].startswith('#cover-'):
                    attachment_id = link['uri'][7:]  # Remove '#cover-'
                    
                    # If we know the position of this attachment's cover page
                    if attachment_id in bookmark_positions:
                        target_page = bookmark_positions[attachment_id]
                        
                        # Create a new goto link
                        new_link = {
                            'kind': fitz.LINK_GOTO,
                            'from': link['from'],
                            'page': target_page,
                            'to': fitz.Point(0, 0),
                            'zoom': 0
                        }
                        
                        # Remove old link and add new one
                        page.delete_link(link)
                        page.insert_link(new_link)
                        links_fixed += 1
    
    except Exception as e:
        print(f"Warning: Error fixing links: {e}")
    
    # Save the merged PDF
    print(f"Saving merged PDF with {links_fixed} fixed links")
    merged_pdf.save(MERGED_PDF)
    final_page_count = len(merged_pdf)
    merged_pdf.close()
    
    print(f"Merged PDF created at: {MERGED_PDF} ({final_page_count} pages)")

def main():
    """
    Main function to run the script.
    """
    # Ensure output directory exists
    os.makedirs('output-files', exist_ok=True)
    
    try:
        # Read data from Excel
        attachments = read_attachment_data()
        
        # Merge PDFs
        merge_pdfs(attachments)
        
    except Exception as e:
        print(f"Error: {e}")
        return 1
    
    return 0

if __name__ == "__main__":
    exit(main()) 