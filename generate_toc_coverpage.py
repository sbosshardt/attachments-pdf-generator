#!/usr/bin/env python3
"""
Generate Table of Contents and Cover Pages PDF

This script reads attachment information from an Excel file and generates
a PDF containing a table of contents and cover pages for each attachment.
"""

import os
import openpyxl
from weasyprint import HTML
from datetime import datetime
import fitz  # PyMuPDF

# File paths
EXCEL_FILE = os.path.join('input-files', 'input-pdfs.xlsx')
TITLE_PAGE = os.path.join('input-files', 'title-page.pdf')
FOREWORD_PAGE = os.path.join('input-files', 'foreword.pdf')
OUTPUT_TOC = os.path.join('output-files', 'toc-coverpage.pdf')
OUTPUT_PDF = os.path.join('output-files', 'weasyoutput.pdf')

def read_attachment_data():
    """
    Read attachment data from Excel file.
    
    Returns:
        list: List of dictionaries containing attachment data
    """
    print(f"Opening Excel file: {EXCEL_FILE}")
    
    # Check if input file exists
    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(f"Excel file not found: {EXCEL_FILE}")
    
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    
    # Check if sheet exists
    if "Attachments Prep" not in workbook.sheetnames:
        raise ValueError(f"Sheet 'Attachments Prep' not found in {EXCEL_FILE}")
    
    sheet = workbook["Attachments Prep"]
    
    # Find headers
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value)
    
    # Field mapping (convert sheet headers to our standardized field names)
    field_mapping = {
        'Attachment Number': ['Attachment Number', 'Attachment #', 'Number'],
        'Title': ['Title', 'Document Title'],
        'Page count': ['Page count', 'Pages', 'Page Count'],
        'Additional Remarks about File': ['Additional Remarks about File', 'Remarks', 'Notes'],
        'Body': ['Body', 'Description', 'Body (Description)'],
        'Filename Reference': ['Filename Reference', 'Filename', 'File'],
        'Date (time Pacific)': ['Date (time Pacific)', 'Date', 'Document Date'],
        'Language': ['Language', 'Lang', 'Language Code'],
        'Category': ['Category', 'Document Category'],
        'Document Type': ['Document Type', 'Type'],
        'Confidentiality': ['Confidentiality', 'Confidential'],
        'Source URL (when available)': ['Source URL (when available)', 'Source URL', 'URL'],
        'Exclude': ['Exclude', 'Skip']
    }
    
    # Map headers to indices
    header_indices = {}
    for field, possible_headers in field_mapping.items():
        for i, header in enumerate(headers):
            if header and any(possible_match.lower() == header.lower() for possible_match in possible_headers):
                header_indices[field] = i
                break
    
    # Print detected headers
    print("Detected headers:")
    for field, idx in header_indices.items():
        print(f"  {field} -> column {idx+1}")
    
    # Check if we found all required fields
    required_fields = ['Attachment Number', 'Title']
    for field in required_fields:
        if field not in header_indices:
            raise ValueError(f"Required field '{field}' not found in Excel headers")
    
    data = []
    # Start from row 2 (skip header)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Skip empty rows
        if not any(row):
            continue
        
        # Skip rows marked as excluded
        exclude_idx = header_indices.get('Exclude')
        if exclude_idx is not None and row[exclude_idx] in (True, 'TRUE', 'True', 'true', 'YES', 'Yes', 'yes', '1', 1):
            continue
        
        attachment = {}
        for field, idx in header_indices.items():
            if field != 'Exclude':  # We've already used this field for filtering
                value = row[idx] if idx < len(row) else None
                
                # Format date if it's a datetime object (remove the 00:00:00 time if it's midnight)
                if field == 'Date (time Pacific)' and value:
                    if isinstance(value, datetime):
                        # Only show the time if it's not midnight
                        if value.hour == 0 and value.minute == 0 and value.second == 0:
                            value = value.strftime('%Y-%m-%d')
                        else:
                            value = value.strftime('%Y-%m-%d %H:%M:%S')
                
                attachment[field] = value
        
        # Set default value for Language if not found
        if 'Language' not in attachment or not attachment['Language']:
            attachment['Language'] = 'EN'
        
        data.append(attachment)
    
    print(f"Found {len(data)} attachments")
    return data

def get_css_styles():
    """
    Return the CSS styles for the HTML document.
    
    Returns:
        str: CSS styles
    """
    return """
        @page {
            size: 8.5in 11in;
            margin: 0.5in 0.5in 0.5in 0.5in;
        }
        
        @page toc-page {
            @bottom-center {
                font-family: Arial, sans-serif;
                font-size: 10pt;
            }
        }
        
        body {
            font-family: Arial, sans-serif;
            font-size: 12pt;
            line-height: 1.5;
            counter-reset: page 1;
        }
        
        h1 {
            font-size: 18pt;
            font-weight: bold;
            text-align: center;
            margin-top: 0.5in;
            margin-bottom: 0.25in;
        }
        
        .toc-container {
            page: toc-page;
            break-after: page;
            width: 100%;
        }
        
        /* TOC Table Styling */
        table.toc-table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 0.25in;
            margin-bottom: 0.25in;
            table-layout: fixed;
        }
        
        /* Column widths */
        table.toc-table td.attachment-num {
            width: 1.5in;
            vertical-align: top;
            padding-bottom: 0.15in;
        }
        
        table.toc-table td.attachment-title {
            vertical-align: top;
            padding-bottom: 0.15in;
        }
        
        table.toc-table td.page-num {
            width: 0.5in;
            text-align: right;
            vertical-align: top;
            padding-bottom: 0.15in;
        }
        
        /* Leader line */
        table.toc-table td.attachment-title {
            position: relative;
        }
        
        table.toc-table td.attachment-title::after {
            content: "";
            position: absolute;
            bottom: 0.5em;
            left: 0;
            width: 100%;
            border-bottom: 1px dotted black;
            z-index: -1;
        }
        
        /* Link styling */
        a.toc-link {
            color: blue;
            text-decoration: underline;
            white-space: nowrap;
        }
        
        /* Cover page styling */
        .cover-page {
            page-break-before: always;
            page-break-after: always;
            text-align: center;
            position: relative;
            height: 10in; /* Page height minus margins */
        }
        
        .cover-number {
            font-size: 14pt;
            margin-top: 0.25in;
            margin-bottom: 0.1in;
        }
        
        .cover-title {
            font-size: 16pt;
            font-weight: bold;
            margin-bottom: 0.25in;
        }
        
        .cover-info {
            position: absolute;
            bottom: 0.5in;
            width: 100%;
            text-align: center;
            font-size: 12pt;
        }
        
        .cover-metadata {
            width: 90%;
            margin: 0.2in auto;
            text-align: left;
            border-collapse: collapse;
        }
        
        .cover-metadata th {
            text-align: right;
            vertical-align: top;
            padding: 0.1in 0.2in 0.1in 0;
            font-weight: bold;
            width: 20%;
            white-space: nowrap;
        }
        
        .cover-metadata td {
            text-align: left;
            vertical-align: top;
            padding: 0.1in 0 0.1in 0;
        }
        
        .cover-description-header {
            width: 90%;
            margin: 0.3in auto 0.1in auto;
            text-align: left;
            font-weight: bold;
        }
        
        .cover-description {
            width: 90%;
            margin: 0 auto 0.3in auto;
            text-align: left;
            font-style: italic;
            line-height: 1.6;
        }
        
        .cover-remarks {
            width: 90%;
            margin: 0.3in auto;
            text-align: left;
            font-size: 11pt;
            border-top: 1px solid #999;
            padding-top: 0.2in;
        }
    """

def normalize_attachment_number(num):
    """
    Normalize attachment number to integer if possible.
    
    Args:
        num: The attachment number
    
    Returns:
        The normalized attachment number
    """
    if isinstance(num, float) and num.is_integer():
        return int(num)
    return num

def normalize_page_count(count):
    """
    Normalize page count to integer if possible.
    
    Args:
        count: The page count
    
    Returns:
        int: The normalized page count
    """
    if isinstance(count, float) and count.is_integer():
        return int(count)
    elif not isinstance(count, int):
        try:
            return int(float(count))
        except (ValueError, TypeError):
            return 1
    return count

def determine_foreword_page_count():
    """
    Determine the number of pages in the foreword document.
    
    Returns:
        int: Number of pages in the foreword
    """
    if os.path.exists(FOREWORD_PAGE):
        foreword_pdf = fitz.open(FOREWORD_PAGE)
        count = len(foreword_pdf)
        foreword_pdf.close()
        return count
    return 0

def calculate_page_map(attachments):
    """
    Calculate page numbers for each attachment.
    
    Args:
        attachments: List of dictionaries containing attachment data
    
    Returns:
        dict: Mapping of attachment numbers to page numbers
    """
    sorted_data = sorted(attachments, key=lambda x: x.get('Attachment Number', 0))
    
    # Start page is calculated based on Title page + TOC pages + first cover
    toc_entries = len(sorted_data)
    toc_pages = max(1, min(3, (toc_entries + 14) // 15))  # Estimate 15 entries per page
    foreword_pages = determine_foreword_page_count()
    start_page = 1 + foreword_pages + toc_pages + 1  # Title page + Foreword + TOC pages + first cover page
    
    # Track page counts for each attachment
    current_page = start_page
    page_map = {}
    
    for attachment in sorted_data:
        attachment_num = normalize_attachment_number(attachment.get('Attachment Number', ''))
        
        # Store the page number for this attachment
        page_map[str(attachment_num)] = current_page
        
        # Get the number of content pages for this attachment
        page_count = normalize_page_count(attachment.get('Page count', 1))
        
        # Move to the next cover page
        current_page += page_count + 1  # Add content pages plus next cover
    
    return page_map

def generate_toc_html(sorted_data, page_map):
    """
    Generate HTML for the table of contents.
    
    Args:
        sorted_data: Sorted list of attachment data
        page_map: Mapping of attachment numbers to page numbers
    
    Returns:
        str: HTML for the table of contents
    """
    html = """
    <div class="toc-container">
        <h1>Table of Contents</h1>
        <table class="toc-table">
    """
    
    # Add a TOC entry for each attachment
    for attachment in sorted_data:
        attachment_num = normalize_attachment_number(attachment.get('Attachment Number', ''))
        title = attachment.get('Title', 'Untitled')
        
        # Get the page number for this attachment
        page_number = page_map.get(str(attachment_num), 0)
        
        # Add the TOC entry with table structure
        html += f"""
        <tr id="toc-entry-{attachment_num}">
            <td class="attachment-num">
                <a class="toc-link" href="#cover-{attachment_num}">Attachment {attachment_num}</a>
            </td>
            <td class="attachment-title">{title}</td>
            <td class="page-num">{page_number}</td>
        </tr>
        """
    
    # Close the TOC section
    html += """
        </table>
    </div>
    """
    
    return html

def generate_cover_page_html(attachment, page_number):
    """
    Generate HTML for a single cover page.
    
    Args:
        attachment: Dictionary containing attachment data
        page_number: The page number to display on the cover page
    
    Returns:
        str: HTML for the cover page
    """
    attachment_num = normalize_attachment_number(attachment.get('Attachment Number', ''))
    title = attachment.get('Title', 'Untitled')
    
    # Extract additional fields (if available)
    date = attachment.get('Date (time Pacific)', '')
    category = attachment.get('Category', '')
    document_type = attachment.get('Document Type', '')
    page_count = attachment.get('Page count', '')
    confidentiality = attachment.get('Confidentiality', '')
    body = attachment.get('Body', '')
    remarks = attachment.get('Additional Remarks about File', '')
    source_url = attachment.get('Source URL (when available)', '')
    
    html = f"""
    <div class="cover-page" id="cover-{attachment_num}">
        <div class="cover-number">Attachment {attachment_num}</div>
        <div class="cover-title">{title}</div>
        
        <table class="cover-metadata">
    """
    
    # Only add fields that have values
    if date:
        html += f"""
        <tr>
            <th>Date (time Pacific):</th>
            <td>{date}</td>
        </tr>
        """
        
    if category:
        html += f"""
        <tr>
            <th>Category:</th>
            <td>{category}</td>
        </tr>
        """
        
    if document_type:
        html += f"""
        <tr>
            <th>Document Type:</th>
            <td>{document_type}</td>
        </tr>
        """
        
    if page_count:
        html += f"""
        <tr>
            <th>Page Count:</th>
            <td>{page_count}</td>
        </tr>
        """
        
    if confidentiality:
        html += f"""
        <tr>
            <th>Confidentiality:</th>
            <td>{confidentiality}</td>
        </tr>
        """
        
    if source_url:
        html += f"""
        <tr>
            <th>Source:</th>
            <td>{source_url}</td>
        </tr>
        """
        
    html += """
        </table>
    """
    
    # Add body/description if available
    if body:
        html += f"""
        <div class="cover-description-header">Description:</div>
        <div class="cover-description">
            {body}
        </div>
        """
        
    # Add remarks if available
    if remarks:
        html += f"""
        <div class="cover-remarks">
            <strong>Additional Remarks:</strong><br>
            {remarks}
        </div>
        """
    
    # Add page number at the bottom
    html += f"""
        <div class="cover-info">
            Page {page_number}
        </div>
    </div>
    """
    
    return html

def generate_html(data):
    """
    Generate HTML content for the PDF.
    
    Args:
        data: List of dictionaries containing attachment data
        
    Returns:
        String containing HTML content
    """
    # Sort the data by attachment number
    sorted_data = sorted(data, key=lambda x: x.get('Attachment Number', 0))
    
    # Calculate page numbers for each attachment
    page_map = calculate_page_map(sorted_data)
    
    # Start building the HTML document
    html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Table of Contents and Cover Pages</title>
    <style>{get_css_styles()}</style>
</head>
<body>
    """
    
    # Add Table of Contents
    html += generate_toc_html(sorted_data, page_map)
    
    # Add cover pages for each attachment
    for attachment in sorted_data:
        attachment_num = normalize_attachment_number(attachment.get('Attachment Number', ''))
        page_number = page_map.get(str(attachment_num), 0)
        html += generate_cover_page_html(attachment, page_number)
    
    # Close the HTML document
    html += """
</body>
</html>
    """
    
    # Save the HTML to a file for debugging
    with open('output-files/toc-debug.html', 'w') as f:
        f.write(html)
    print("Saved HTML to output-files/toc-debug.html for debugging")
    
    return html

def main():
    """
    Main function to run the script.
    """
    # Ensure output directory exists
    os.makedirs('output-files', exist_ok=True)
    
    try:
        # Read data from Excel
        attachments = read_attachment_data()
        
        # Print attachment page counts for debugging
        print("\nDEBUG - Attachment page counts:")
        total_content_pages = 0
        
        # First count actual page counts
        attachment_pages = {}
        for attachment in attachments:
            attachment_num = attachment.get('Attachment Number', '')
            if isinstance(attachment_num, float) and attachment_num.is_integer():
                attachment_num = int(attachment_num)
            
            page_count = attachment.get('Page count', 1)
            if isinstance(page_count, float) and page_count.is_integer():
                page_count = int(page_count)
            elif not isinstance(page_count, int):
                try:
                    page_count = int(float(page_count))
                except (ValueError, TypeError):
                    page_count = 1
                    
            # Store page count and add to total
            attachment_pages[str(attachment_num)] = page_count
            total_content_pages += page_count  # Just the content pages
            print(f"Attachment {attachment_num}: {page_count} pages of content")
        
        # Calculate total pages including cover pages
        total_attachment_pages = total_content_pages + len(attachments)  # Content + cover pages
        
        # Estimate TOC pages based on number of entries
        toc_entries = len(attachments)
        toc_pages = max(1, min(3, (toc_entries + 14) // 15))  # Estimate 15 entries per page
        
        # Check if foreword exists
        foreword_pages = determine_foreword_page_count()
        
        # Total estimated pages: Title page + Foreword + TOC pages + attachment pages (content + covers)
        estimated_total = 1 + foreword_pages + toc_pages + total_attachment_pages
        
        print(f"\nPage count summary:")
        print(f"- Title page: 1 page")
        if foreword_pages > 0:
            print(f"- Foreword: {foreword_pages} page(s)")
        print(f"- Table of Contents: approximately {toc_pages} pages ({toc_entries} entries)")
        print(f"- Cover pages: {len(attachments)} pages (one per attachment)")
        print(f"- Attachment content: {total_content_pages} pages")
        print(f"Estimated total pages: {estimated_total}")
        
        # Now verify our page number calculations
        print("\nDEBUG - Page numbering verification:")
        
        # Calculate expected page numbers
        start_page = 1 + foreword_pages + toc_pages + 1  # Title page + Foreword + TOC pages + first cover
        current_page = start_page
        
        # Page map to use for both TOC display and bookmarks
        page_map = {}
        
        for attachment in attachments:
            attachment_num = attachment.get('Attachment Number', '')
            if isinstance(attachment_num, float) and attachment_num.is_integer():
                attachment_num = int(attachment_num)
                
            page_map[str(attachment_num)] = current_page
            print(f"Attachment {attachment_num} cover page should be on page {current_page}")
            
            # Move to next cover page
            page_count = attachment_pages.get(str(attachment_num), 1)
            current_page += page_count + 1  # Current content + next cover
        
        # Generate HTML 
        html_content = generate_html(attachments)
        
        # Convert HTML to PDF using WeasyPrint
        print(f"\nGenerating TOC and cover pages: {OUTPUT_TOC}")
        HTML(string=html_content).write_pdf(OUTPUT_TOC)
        
        # Check TOC page count
        toc_doc = fitz.open(OUTPUT_TOC)
        actual_toc_pages = len(toc_doc)
        
        # Check for links in the TOC PDF
        print("\nDEBUG - Checking links in TOC PDF:")
        toc_links_found = 0
        for page_num in range(toc_doc.page_count):
            page = toc_doc[page_num]
            links = page.get_links()
            if links:
                toc_links_found += len(links)
                print(f"Found {len(links)} links on page {page_num+1}")
                for idx, link in enumerate(links):
                    print(f"  Link {idx+1}: {link}")
        
        if toc_links_found == 0:
            print("WARNING: No links found in TOC PDF. Adding links manually...")
            
            # Find TOC page 
            toc_page_idx = -1
            for page_idx in range(toc_doc.page_count):
                page = toc_doc[page_idx]
                text = page.get_text()
                if "Table of Contents" in text:
                    toc_page_idx = page_idx
                    print(f"Found TOC on page {page_idx+1}")
                    break
            
            if toc_page_idx >= 0:
                # For each attachment, find text "Attachment X" on TOC page and create a link to the cover page
                toc_page = toc_doc[toc_page_idx]
                links_added = 0
                
                for attachment in attachments:
                    attachment_num = attachment.get('Attachment Number', '')
                    if isinstance(attachment_num, float) and attachment_num.is_integer():
                        attachment_num = int(attachment_num)
                    
                    # Find text on the page
                    search_text = f"Attachment {attachment_num}"
                    rects = toc_page.search_for(search_text)
                    
                    if rects:
                        # The first occurrence is the one in the TOC
                        rect = rects[0]
                        
                        # Calculate target page
                        target_page = toc_page_idx + 1 + attachment_num  # Simple heuristic
                        if target_page < toc_doc.page_count:
                            # Create a new internal link
                            link = {
                                "kind": fitz.LINK_GOTO,
                                "from": rect,  # rectangle containing found text
                                "page": target_page,
                                "to": fitz.Point(0, 0),  # top of target page
                                "zoom": 0,  # default zoom
                            }
                            
                            toc_page.insert_link(link)
                            links_added += 1
                            print(f"Added link for {search_text} pointing to page {target_page+1}")
                
                if links_added > 0:
                    print(f"Added {links_added} links to the TOC")
                    toc_doc.save(OUTPUT_TOC)
        
        toc_doc.close()
        print(f"Actual TOC PDF page count: {actual_toc_pages}")
        
        # Check if title page exists
        if os.path.exists(TITLE_PAGE):
            print(f"Adding title page from: {TITLE_PAGE}")
            
            # Create a new PDF with title page followed by TOC and cover pages
            merged_pdf = fitz.open()
            
            # Add the title page
            title_pdf = fitz.open(TITLE_PAGE)
            title_page_count = len(title_pdf)
            merged_pdf.insert_pdf(title_pdf)
            print(f"Title page has {title_page_count} page(s)")
            
            # Add the foreword if it exists
            foreword_page_offset = title_page_count
            if os.path.exists(FOREWORD_PAGE):
                print(f"Adding foreword from: {FOREWORD_PAGE}")
                foreword_pdf = fitz.open(FOREWORD_PAGE)
                foreword_page_count = len(foreword_pdf)
                merged_pdf.insert_pdf(foreword_pdf)
                print(f"Foreword has {foreword_page_count} page(s)")
                foreword_page_offset = title_page_count
            else:
                foreword_page_count = 0
            
            # Add the TOC and cover pages
            toc_pdf = fitz.open(OUTPUT_TOC)
            toc_page_count = len(toc_pdf)
            merged_pdf.insert_pdf(toc_pdf)
            
            # Create bookmarks for the PDF
            toc = []
            
            # Add title page bookmark
            toc.append([1, "Title Page", 0])
            
            # Add foreword bookmark if it exists
            if os.path.exists(FOREWORD_PAGE):
                toc.append([1, "Foreword", foreword_page_offset])
            
            # Find TOC page in the merged document
            toc_page_idx = title_page_count + foreword_page_count
            toc.append([1, "Table of Contents", toc_page_idx])
            
            # Map actual cover page indices in the merged document
            actual_cover_pages = {}
            
            # Find all cover pages by text pattern
            for page_num in range(merged_pdf.page_count):
                page = merged_pdf[page_num]
                text = page.get_text()
                
                for attachment_num in attachment_pages.keys():
                    if f"Attachment {attachment_num}" in text:
                        # Make sure this isn't just a mention in another cover page
                        if "Page" in text and f"Attachment {attachment_num}" in text.split('\n')[:5]:
                            actual_cover_pages[attachment_num] = page_num
                            # Add bookmark for the attachment
                            attachment_data = next((a for a in attachments if str(a.get('Attachment Number', '')) == attachment_num), None)
                            title = attachment_data.get('Title', 'Untitled') if attachment_data else 'Untitled'
                            toc.append([1, f"Attachment {attachment_num}: {title}", page_num])
                            print(f"Added bookmark for Attachment {attachment_num} on page {page_num+1}")
                            break
            
            # Set bookmarks
            merged_pdf.set_toc(toc)
            print(f"Created {len(toc)} bookmarks")
            
            # Save merged PDF
            merged_pdf.save(OUTPUT_PDF)
            merged_pdf.close()
            
            print(f"PDF generated at: {OUTPUT_PDF}")
            
            # Clean up
            title_pdf.close()
            toc_pdf.close()
            
            # Try to fix links in the final PDF
            try:
                print("\nFixing links in final PDF...")
                final_pdf = fitz.open(OUTPUT_PDF)
                final_page_count = len(final_pdf)
                
                # Find TOC page by content
                toc_page_idx = -1
                for page_idx in range(final_page_count):
                    page = final_pdf[page_idx]
                    text = page.get_text()
                    if "Table of Contents" in text:
                        toc_page_idx = page_idx
                        print(f"Found TOC on page {page_idx+1}")
                        break
                
                # Fix links
                links_fixed = 0
                if toc_page_idx >= 0:
                    print(f"Examining links on TOC page {toc_page_idx+1}")
                    toc_page = final_pdf[toc_page_idx]
                    toc_links = toc_page.get_links()
                    print(f"Found {len(toc_links)} links on TOC page")
                    
                    for link in toc_links:
                        if 'uri' in link and link['uri'].startswith('#cover-'):
                            attachment_id = link['uri'][7:]  # Remove '#cover-'
                            print(f"Found TOC link to Attachment {attachment_id}")
                            # If we have a page number for this attachment in our map
                            if attachment_id in actual_cover_pages:
                                target_page = actual_cover_pages[attachment_id]
                                print(f"  Fixing link to point to page {target_page+1}")
                                # Validate the target page is within the document
                                if target_page >= 0 and target_page < final_page_count:
                                    # Create a new goto link
                                    new_link = {
                                        'kind': fitz.LINK_GOTO,
                                        'from': link['from'],  # rectangle
                                        'page': target_page,   # target page
                                        'to': fitz.Point(0, 0),  # top of page
                                        'zoom': 0  # default zoom
                                    }
                                    # Remove old link and add new one
                                    toc_page.delete_link(link)
                                    toc_page.insert_link(new_link)
                                    links_fixed += 1
                                else:
                                    print(f"  Warning: Target page {target_page} is out of range (0-{final_page_count-1})")
                            else:
                                print(f"  Warning: No mapping found for Attachment {attachment_id}")
                
                # Now scan all pages for any remaining links
                for page_num in range(final_page_count):
                    if page_num == toc_page_idx:
                        continue  # Skip TOC page as we've already processed it
                        
                    page = final_pdf[page_num]
                    links = page.get_links()
                    
                    for link in links:
                        # If this is a named destination link (uri starts with #)
                        if 'uri' in link and link['uri'].startswith('#cover-'):
                            attachment_id = link['uri'][7:]  # Remove '#cover-'
                            print(f"Found link to Attachment {attachment_id} on page {page_num+1}")
                            
                            # If we have a page number for this attachment
                            if attachment_id in actual_cover_pages:
                                # Get the page number from our mapped locations
                                target_page = actual_cover_pages[attachment_id]
                                print(f"  Fixing link to point to page {target_page+1}")
                                
                                # Validate the target page is within the document
                                if target_page >= 0 and target_page < final_page_count:
                                    # Create a new page link
                                    new_link = {
                                        'kind': fitz.LINK_GOTO,
                                        'from': link['from'],
                                        'page': target_page,
                                        'to': fitz.Point(0, 0),
                                        'zoom': 0
                                    }
                                    # Remove old link and add new one
                                    page.delete_link(link)
                                    page.insert_link(new_link)
                                    links_fixed += 1
                                else:
                                    print(f"  Warning: Target page {target_page} is out of range (0-{final_page_count-1})")
                            else:
                                print(f"  Warning: No mapping found for Attachment {attachment_id}")
                
                if links_fixed > 0:
                    print(f"Fixed {links_fixed} links in the document")
                    # Save again if links were fixed
                    final_pdf.save(OUTPUT_PDF)
                else:
                    print("No links needed fixing")
            except Exception as e:
                print(f"Warning: Error fixing links: {e}")
                print("Proceeding with original links")
                import traceback
                traceback.print_exc()
            
            final_pdf.close()
            
        else:
            # If no title page, just use the TOC PDF as the output
            print(f"Title page not found at {TITLE_PAGE}, using TOC only")
            import shutil
            shutil.copy(OUTPUT_TOC, OUTPUT_PDF)
            print(f"PDF generated at: {OUTPUT_PDF}")
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return 1
    
    return 0

if __name__ == "__main__":
    exit(main()) 